import contextlib
import shutil
import uuid
from pathlib import Path

from openpyxl import load_workbook

from app.config import settings
from app.models import ImportResult, TransactionCreate
from app.models.import_models import (
    ExistingDbInfo,
    FilePreviewResponse,
    ImportExecuteResponse,
    ImportRequest,
    SheetImportConfig,
    SheetImportResult,
    SheetInfo,
    SheetPreview,
)
from app.services.db_service import db_service
from app.utils.excel_parser import COLUMN_MAPPING, parse_excel_file, parse_sheet


class ImportService:
    """Service for importing Excel files into databases."""

    def __init__(self):
        self._temp_dir = settings.db_path / "temp"
        self._temp_dir.mkdir(parents=True, exist_ok=True)

    def preview_excel(self, file_path: str | Path) -> FilePreviewResponse:
        """
        Parse Excel file and return preview information for the wizard.
        Stores file temporarily with a file_id for later execution.
        """
        file_path = Path(file_path)
        file_id = str(uuid.uuid4())

        # Store file temporarily
        temp_file = self._temp_dir / f"{file_id}.xlsx"
        shutil.copy(file_path, temp_file)

        wb = load_workbook(filename=file_path, data_only=True)
        sheets: list[SheetPreview] = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                continue

            # Find headers and mapping
            headers: list[str] = []
            detected_mapping: dict[str, str] = {}
            header_row_idx = None

            for i, row in enumerate(rows):
                if row is None:
                    continue
                for _col_idx, cell in enumerate(row):
                    if cell is None:
                        continue
                    cell_str = str(cell).strip()
                    if cell_str in COLUMN_MAPPING:
                        headers.append(cell_str)
                        detected_mapping[cell_str] = COLUMN_MAPPING[cell_str]
                if headers:
                    header_row_idx = i
                    break

            if not headers:
                continue

            # Count data rows
            data_rows = rows[header_row_idx + 1 :] if header_row_idx is not None else []
            row_count = sum(1 for row in data_rows if row and not all(cell is None for cell in row))

            # Get sample rows (first 3)
            sample_rows: list[dict] = []
            for row in data_rows[:3]:
                if row and not all(cell is None for cell in row):
                    sample = {}
                    for col_idx, header in enumerate(headers):
                        if col_idx < len(row):
                            val = row[col_idx]
                            sample[header] = str(val) if val is not None else ""
                    sample_rows.append(sample)

            sheet_info = SheetInfo(
                name=sheet_name,
                row_count=row_count,
                headers=headers,
                detected_mapping=detected_mapping,
                sample_rows=sample_rows,
            )

            # Check for existing database
            existing_db = self._get_existing_db_info(sheet_name)

            sheets.append(SheetPreview(sheet=sheet_info, existing_db=existing_db))

        return FilePreviewResponse(file_id=file_id, sheets=sheets)

    def _get_existing_db_info(self, account_name: str) -> ExistingDbInfo | None:
        """Get information about existing database for an account."""
        account = db_service.get_account_by_name(account_name)
        if not account:
            return None

        summary = db_service.get_account_summary(account.id)
        return ExistingDbInfo(
            account_id=account.id,
            account_name=account.name,
            existing_row_count=summary.transaction_count if summary else 0,
            last_transaction_date=summary.last_transaction_date if summary else None,
        )

    def execute_import(self, request: ImportRequest) -> ImportExecuteResponse:
        """Execute import with user's selections."""
        temp_file = self._temp_dir / f"{request.file_id}.xlsx"

        if not temp_file.exists():
            return ImportExecuteResponse(
                success=False,
                results=[],
                total_rows_imported=0,
                total_rows_skipped=0,
            )

        wb = load_workbook(filename=temp_file, data_only=True)
        results: list[SheetImportResult] = []
        total_imported = 0
        total_skipped = 0

        for config in request.sheets:
            if not config.selected:
                continue

            result = self._import_sheet(wb, config)
            results.append(result)
            total_imported += result.rows_imported
            total_skipped += result.rows_skipped

        # Cleanup temp file
        with contextlib.suppress(Exception):
            temp_file.unlink()

        return ImportExecuteResponse(
            success=all(r.success for r in results),
            results=results,
            total_rows_imported=total_imported,
            total_rows_skipped=total_skipped,
        )

    def _import_sheet(self, wb, config: SheetImportConfig) -> SheetImportResult:
        """Import a single sheet based on user configuration."""
        # Use target_db_name for database lookup/creation, target_account_name for display
        db_name = config.target_db_name or config.sheet_name
        account_name = config.target_account_name or config.sheet_name

        try:
            if config.sheet_name not in wb.sheetnames:
                return SheetImportResult(
                    sheet_name=config.sheet_name,
                    account_name=account_name,
                    success=False,
                    rows_imported=0,
                    rows_skipped=0,
                    error=f"Sheet '{config.sheet_name}' not found",
                )

            sheet = wb[config.sheet_name]
            transactions_data = parse_sheet(sheet)

            # Get or create account using the target db name for storage, account name for display
            account = db_service.get_account_by_db_path(db_name)
            is_new_account = account is None

            if account is None:
                account = db_service.create_account(account_name, db_name)

            rows_imported = 0
            rows_skipped = 0

            if config.import_mode == "override":
                # Clear existing transactions
                if not is_new_account:
                    db_service.clear_account_transactions(account.id)

                # Import all transactions
                transactions = self._convert_to_transactions(account.id, transactions_data)
                rows_imported = db_service.insert_transactions(account.id, transactions)
            else:
                # Append mode: only add rows newer than last transaction
                summary = db_service.get_account_summary(account.id)
                last_date = summary.last_transaction_date if summary else None

                transactions = []
                for txn_data in transactions_data:
                    txn_date = txn_data.get("date")
                    if last_date is None or (txn_date and txn_date > last_date):
                        transactions.append(self._create_transaction(account.id, txn_data))
                        rows_imported += 1
                    else:
                        rows_skipped += 1

                if transactions:
                    db_service.insert_transactions(account.id, transactions)

            return SheetImportResult(
                sheet_name=config.sheet_name,
                account_name=account.name,
                success=True,
                rows_imported=rows_imported,
                rows_skipped=rows_skipped,
            )

        except Exception as e:
            return SheetImportResult(
                sheet_name=config.sheet_name,
                account_name=account_name,
                success=False,
                rows_imported=0,
                rows_skipped=0,
                error=str(e),
            )

    def _convert_to_transactions(
        self, account_id: str, transactions_data: list[dict]
    ) -> list[TransactionCreate]:
        """Convert parsed transaction data to TransactionCreate objects."""
        return [self._create_transaction(account_id, txn) for txn in transactions_data]

    def _create_transaction(self, account_id: str, txn: dict) -> TransactionCreate:
        """Create a TransactionCreate from parsed data."""
        return TransactionCreate(
            account_id=account_id,
            date=txn["date"],
            description=txn["description"],
            payment_method=txn.get("payment_method"),
            category=txn.get("category", "לא מסווג"),
            details=txn.get("details"),
            reference=txn.get("reference"),
            expense=txn.get("expense", 0.0),
            income=txn.get("income", 0.0),
            balance=txn.get("balance", 0.0),
            raw_date_string=txn.get("raw_date_string", ""),
        )

    def import_excel(self, file_path: str | Path) -> ImportResult:
        """
        Import an Excel file with multiple sheets.
        Each sheet becomes an account with its own transactions.
        """
        errors = []
        accounts_created = 0
        transactions_imported = 0
        last_transaction_date: int | None = None

        try:
            sheets_data = parse_excel_file(file_path)
        except Exception as e:
            return ImportResult(
                accounts_created=0,
                transactions_imported=0,
                errors=[f"Failed to parse Excel file: {str(e)}"],
            )

        for sheet_name, transactions_data in sheets_data.items():
            try:
                # Get or create account
                account = db_service.get_account_by_name(sheet_name)
                if account:
                    # Clear existing transactions for re-import
                    db_service.clear_account_transactions(account.id)
                else:
                    account = db_service.create_account(sheet_name)
                    accounts_created += 1

                # Convert to TransactionCreate objects
                transactions = [
                    TransactionCreate(
                        account_id=account.id,
                        date=txn["date"],
                        description=txn["description"],
                        payment_method=txn.get("payment_method"),
                        category=txn.get("category", "לא מסווג"),
                        details=txn.get("details"),
                        reference=txn.get("reference"),
                        expense=txn.get("expense", 0.0),
                        income=txn.get("income", 0.0),
                        balance=txn.get("balance", 0.0),
                        raw_date_string=txn.get("raw_date_string", ""),
                    )
                    for txn in transactions_data
                ]

                # Insert transactions
                count = db_service.insert_transactions(account.id, transactions)
                transactions_imported += count

                # Track last transaction date
                for txn in transactions:
                    if txn.date and (
                        last_transaction_date is None or txn.date > last_transaction_date
                    ):
                        last_transaction_date = txn.date

            except Exception as e:
                errors.append(f"Error importing sheet '{sheet_name}': {str(e)}")

        return ImportResult(
            accounts_created=accounts_created,
            transactions_imported=transactions_imported,
            last_transaction_date=last_transaction_date,
            errors=errors,
        )


import_service = ImportService()
