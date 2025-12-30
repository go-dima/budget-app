import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# Hebrew column name to field mapping
COLUMN_MAPPING = {
    "תאריך": "date",
    "תיאור": "description",
    "אמצעי תשלום": "payment_method",
    "קטגוריה": "category",
    "פירוט": "details",
    "אסמכתא": "reference",
    "חובה": "expense",
    "זכות": "income",
    "יתרה": "balance",
}

# Hebrew month names for parsing dates
HEBREW_MONTHS = {
    "ינו": 1,
    "ינואר": 1,
    "פבר": 2,
    "פברואר": 2,
    "מרץ": 3,
    "מרס": 3,
    "אפר": 4,
    "אפריל": 4,
    "מאי": 5,
    "יונ": 6,
    "יוני": 6,
    "יול": 7,
    "יולי": 7,
    "אוג": 8,
    "אוגוסט": 8,
    "ספט": 9,
    "ספטמבר": 9,
    "אוק": 10,
    "אוקטובר": 10,
    "נוב": 11,
    "נובמבר": 11,
    "דצמ": 12,
    "דצמבר": 12,
}

# English month abbreviations
ENGLISH_MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def parse_date(date_value) -> tuple[int | None, str]:
    """
    Parse date from various formats.
    Returns (unix_timestamp, raw_date_string).
    """
    if date_value is None:
        return None, ""

    raw_string = str(date_value).strip()

    # If it's already a datetime object
    if isinstance(date_value, datetime):
        return int(date_value.timestamp()), raw_string

    # Try to parse string formats
    date_str = raw_string.lower()

    # Try DD/MM/YYYY or DD-MM-YYYY
    match = re.match(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})", date_str)
    if match:
        day, month, year = int(match.group(1)), int(match.group(2)), int(match.group(3))
        if year < 100:
            year += 2000
        try:
            dt = datetime(year, month, day)
            return int(dt.timestamp()), raw_string
        except ValueError:
            pass

    # Try "DD Month YYYY" format (Hebrew or English)
    match = re.match(r"(\d{1,2})\s+([א-ת\w]+)\s+(\d{2,4})", date_str)
    if match:
        day = int(match.group(1))
        month_str = match.group(2).lower()
        year = int(match.group(3))
        if year < 100:
            year += 2000

        month = HEBREW_MONTHS.get(month_str) or ENGLISH_MONTHS.get(month_str)
        if month:
            try:
                dt = datetime(year, month, day)
                return int(dt.timestamp()), raw_string
            except ValueError:
                pass

    return None, raw_string


def parse_amount(value) -> float:
    """Parse amount from various formats (handles commas, spaces, etc)."""
    if value is None:
        return 0.0

    if isinstance(value, int | float):
        return float(value)

    # Remove quotes, commas, spaces
    cleaned = str(value).replace('"', "").replace(",", "").replace(" ", "").strip()

    if not cleaned or cleaned == "-":
        return 0.0

    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def parse_excel_file(file_path: str | Path) -> dict[str, list[dict]]:
    """
    Parse an Excel file with multiple sheets.
    Each sheet represents an account.
    Returns dict with sheet name as key and list of transaction dicts as value.
    """
    wb = load_workbook(filename=file_path, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        transactions = parse_sheet(sheet)
        if transactions:
            result[sheet_name] = transactions

    return result


def parse_sheet(sheet) -> list[dict]:
    """Parse a single Excel sheet into transaction dictionaries."""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row by looking for Hebrew column names
    header_row_idx = None
    column_indices = {}

    for i, row in enumerate(rows):
        if row is None:
            continue

        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_str = str(cell).strip()
            if cell_str in COLUMN_MAPPING:
                column_indices[col_idx] = COLUMN_MAPPING[cell_str]

        if column_indices:
            header_row_idx = i
            break

    if header_row_idx is None or not column_indices:
        return []

    # Parse data rows
    transactions = []
    for row in rows[header_row_idx + 1 :]:
        if row is None or all(cell is None for cell in row):
            continue

        txn = {}
        for col_idx, field_name in column_indices.items():
            if col_idx < len(row):
                value = row[col_idx]

                if field_name == "date":
                    timestamp, raw_string = parse_date(value)
                    txn["date"] = timestamp
                    txn["raw_date_string"] = raw_string
                elif field_name in ("expense", "income", "balance"):
                    txn[field_name] = parse_amount(value)
                else:
                    txn[field_name] = str(value).strip() if value else None

        # Skip rows without valid date or description
        if txn.get("date") is None or not txn.get("description"):
            continue

        # Set defaults for missing fields
        txn.setdefault("payment_method", None)
        txn.setdefault("category", "לא מסווג")  # "Uncategorized" in Hebrew
        txn.setdefault("details", None)
        txn.setdefault("reference", None)
        txn.setdefault("expense", 0.0)
        txn.setdefault("income", 0.0)
        txn.setdefault("balance", 0.0)

        transactions.append(txn)

    return transactions
